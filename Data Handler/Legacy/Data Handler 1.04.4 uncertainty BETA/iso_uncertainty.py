"""
iso_uncertainty.py

Defines a feature that checks newly appended DataFrame rows for a certain keyword
in a chosen column, then appends them to an XLSX file with one sheet per machine_name.
Never rotated daily, stays persistent.
"""

import os
import logging
import threading
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
from settings_manager import SettingsManager

LOG_LOCK = threading.Lock()
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

def log_message(level, message):
    with LOG_LOCK:
        if level == 'info':
            logger.info(message)
        elif level == 'warning':
            logger.warning(message)
        elif level == 'error':
            logger.error(message)
        else:
            logger.debug(message)

def check_iso_uncertainty(df_new_chunk, machine_name):
    """
    Instead of re-reading the entire CSV, we just get the partial DataFrame (df_new_chunk)
    that was newly appended. Then we look for the 'keyword' in the chosen 'column_name'
    for these newly added rows. If found, we copy them to an XLSX file.

    :param df_new_chunk: The partial DataFrame that was appended to the master CSV
    :param machine_name: The parser's machine_name (used as the sheet name)
    """
    # 1) Load global "iso_uncertainty" config from settings
    settings_file = os.path.join(
        os.path.expanduser("~"),
        "csv_parser_configs",
        "settings.json"
    )
    manager = SettingsManager(settings_file)
    iso_config = manager.get_setting("iso_uncertainty", {})

    if not iso_config.get("enabled", False):
        return

    keyword = iso_config.get("keyword", "").strip()
    column_name = iso_config.get("column_name", "").strip()
    output_dir = iso_config.get("output_directory", "").strip()

    if not keyword or not column_name or not output_dir:
        log_message('warning', "ISO Uncertainty settings incomplete; skipping.")
        return

    if not os.path.isdir(output_dir):
        log_message('warning', f"ISO output directory invalid: {output_dir}")
        return

    if df_new_chunk.empty:
        log_message('info', "No new rows in partial DataFrame to check for ISO uncertainty.")
        return

    if column_name not in df_new_chunk.columns:
        log_message('warning', f"Column '{column_name}' not in partial DataFrame. Skipping.")
        return

    # 2) Filter only new rows that match the keyword
    matches = df_new_chunk[df_new_chunk[column_name].astype(str).str.contains(keyword, case=False, na=False)]
    if matches.empty:
        return  # nothing found

    # 3) Append matches to an XLSX file "ISO_Uncertainty.xlsx" in output_dir
    iso_xlsx = os.path.join(output_dir, "ISO_Uncertainty.xlsx")

    if os.path.exists(iso_xlsx):
        wb = load_workbook(iso_xlsx)
    else:
        wb = Workbook()

    sheet_name = machine_name[:31]  # max 31 chars for an Excel sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # remove default "Sheet" if unused
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1 and not any(wb["Sheet"].values):
        wb.remove(wb["Sheet"])

    # 4) Append row by row
    for idx, row_data in matches.iterrows():
        row_list = row_data.astype(str).tolist()
        ws.append(row_list)

    # 5) Save
    try:
        wb.save(iso_xlsx)
        log_message('info', f"Appended {len(matches)} row(s) to ISO XLSX: {iso_xlsx}")
    except Exception as e:
        log_message('error', f"Failed to save ISO XLSX '{iso_xlsx}': {e}")
